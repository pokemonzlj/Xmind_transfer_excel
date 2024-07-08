import openpyxl
from openpyxl.styles import Alignment
from xmindparser import xmind_to_dict
import tkinter as tk
from tkinter import filedialog
from datetime import datetime

"""适用于禅道的用例导入，xmind形式编写用例后，自动生成禅道导入格式的用例"""

def resolve_path(d, lists, title, priority, preconditions, steps, expected):
    """遍历解析 xmind 树形结构，并将解析后的用例数据存储到列表 lists 中"""
    title = title.strip()
    if len(title) == 0:
        concat_title = d['title'].strip()
    else:
        if 'makers' in d.keys():
            level = ''
            if 'priority-2' in d['makers']:
                level = 'P1'
            elif 'priority-1' in d['makers']:
                level = 'P0'
            elif 'priority-3' in d['makers']:
                level = 'P2'
            concat_title = title + '\t' + d['title'].strip() + '\t' + level
        else:
            concat_title = title + '\t' + d['title'].strip()
    if 'topics' not in d:
        lists.append((concat_title, priority, preconditions, steps, expected))
    else:
        for sub_d in d['topics']:
            sub_priority = priority
            sub_preconditions = preconditions
            sub_steps = steps
            sub_expected = expected
            if 'topics' in sub_d:
                sub_priority = ''
            elif sub_d['title'].startswith('前置条件：'):
                sub_preconditions = sub_d['title'].strip().lstrip('前置条件：')
            elif sub_d['title'].startswith('步骤：'):
                sub_steps = sub_d['title'].strip().lstrip('步骤：')
            elif sub_d['title'].startswith('预期：'):
                sub_expected = sub_d['title'].strip().lstrip('预期：')
            resolve_path(sub_d, lists, concat_title, sub_priority, sub_preconditions, sub_steps, sub_expected)
    print(lists)


def xmind_cat(lst):
    wb = openpyxl.Workbook()
    sheetname = wb.sheetnames
    sheet = wb[sheetname[0]]
    title_list = ['所属模块', '用例标题', '优先级', '前置条件', '步骤', '预期', '用例类型']
    for i in range(1, len(title_list) + 1):
        sheet.cell(row=1, column=i).value = title_list[i - 1]
    index = 1
    for h in range(len(lst)):
        lists = []
        resolve_path(lst[h], lists, '', '', '', '', '')
        prev_module_details = ''
        for j in range(len(lists)):
            title, priority, preconditions, steps, expected = lists[j]
            title = title.split('\t')
            start_column = 1
            if 6 > len(title) >= 4:  # 5层或4层结构，对应没有最底层或者没有用例等级
                module_details = '-'.join(lists[j][1:4])
                sheet.cell(row=j + index + 1, column=1).value = lists[j][0]
                sheet.cell(row=j + index + 1, column=2).value = module_details
                if len(lists[j]) > 4:
                    sheet.cell(row=j + index + 1, column=3).value = lists[j][4]
                sheet.cell(row=j + index + 1, column=7).value = "功能测试"
            elif len(title) == 6:
                module_details = '-'.join(lists[j][2:5])
                sheet.cell(row=j + index + 1, column=1).value = lists[j][1]
                sheet.cell(row=j + index + 1, column=2).value = module_details
                sheet.cell(row=j + index + 1, column=3).value = lists[j][5]
                sheet.cell(row=j + index + 1, column=7).value = "功能测试"
            elif len(title) == 7:
                module_details = '-'.join(title[2:5])
                if module_details == prev_module_details:
                    sheet.cell(row=j + index, column=4).value += '\n' + preconditions
                    sheet.cell(row=j + index, column=5).value += '\n' + steps
                    sheet.cell(row=j + index, column=6).value += '\n' + expected
                else:
                    sheet.cell(row=j + index + 1, column=1).value = title[0]
                    sheet.cell(row=j + index + 1, column=2).value = module_details
                    sheet.cell(row=j + index + 1, column=3).value = priority
                    sheet.cell(row=j + index + 1, column=4).value = preconditions
                    sheet.cell(row=j + index + 1, column=5).value = steps
                    sheet.cell(row=j + index + 1, column=6).value = expected
                    sheet.cell(row=j + index + 1, column=7).value = "功能测试"
                    prev_module_details = module_details
        if j == len(lists) - 1:
            index += len(lists)
    now = datetime.now()
    date_time = now.strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"{date_time}_case.xlsx"
    wb.save(filename)


def select_file():
    filepath = filedialog.askopenfilename()
    return filepath


def maintest():
    file_name = select_file()
    out = xmind_to_dict(file_name)
    xmind_cat(out[0]['topic']['topics'])


if __name__ == '__main__':
    maintest()