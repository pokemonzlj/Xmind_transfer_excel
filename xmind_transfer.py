import openpyxl
from openpyxl.styles import Alignment
from xmindparser import xmind_to_dict
import tkinter as tk
from tkinter import filedialog   #文件对话框模块

def resolvePath(dict,lists,title):
    title = title.strip()  # title去除首尾空格
    # print("title is %s, dict is %s"%(title,dict))
    if len(title) == 0:  # 如果title是空字符串，则直接获取value
        concatTitle = dict['title'].strip()
    else:
        if 'makers' in dict.keys():
            level=''
            if 'priority-2' in dict['makers']:
                level='P1'
            elif 'priority-1' in dict['makers']:
                level='P0'
            elif 'priority-3' in dict['makers']:
                level = 'P2'
            concatTitle = title + '\t' + dict['title'].strip() + '\t' + level
        else:
            concatTitle = title + '\t' + dict['title'].strip() #+ '\t' + dict['makers'].strip()
    if dict.__contains__('topics')==False:
        lists.append(concatTitle)
    else:
         for d in dict['topics']:
            resolvePath(d, lists, concatTitle)

def xmind_cat(list ,excelname):
    print(f'当前的list是{list}')
    wb = openpyxl.load_workbook(excelname)
    sheetname = wb.sheetnames
    sheet = wb.get_sheet_by_name(sheetname[0])
    title_list = ['功能模块', '二级模块', '三级模块', 'case详情', 'case等级', '执行结果']
    # 生成第一行中固定表头内容
    for i in range(1, len(title_list)+1):
        sheet.cell(row=1, column=i).value=title_list[i-1]
    # 增量索引
    index = 1    #控制写入的行数
    for h in range(0, len(list)):
        lists = []
        resolvePath(list[h], lists, '')  #把模块切割的list依次放入
        # print('\n'.join(lists))
        print(lists)
        for j in range(0, len(lists)):
            lists[j] = lists[j].split('\t')  #把当前模块下每个场景的内容拆分
            print(lists[j])
            start_column=1
            if len(lists[j])==5 or len(lists[j])==3: #如果有4级case路径
                for info in (lists[j]):
                    sheet.cell(row=j + index + 1, column=start_column).value = info  # 从第二行开始写
                    start_column+=1
            elif len(lists[j])==4:
                for n in range(3):
                    sheet.cell(row=j + index + 1, column=n+1).value =lists[j][n]
                sheet.cell(row=j + index + 1, column=5).value =lists[j][3]   #最后一个等级直接写入第5列
        # 遍历结束lists，给增量索引赋值，跳出for j循环，开始for h循环
        if j == len(lists) - 1:
            index += len(lists)
    max_row = sheet.max_row
    print("max row is %s" % max_row)
    for i in range(1, 4):  #合并头三列的单元格
        start_count = 2  #从第一行开始
        end_count = 2
        for count in range(3, max_row+2): #对照数据从第三行开始,要合并最后一轮，所以不能走到最后一个有数据的格，还得再往下走一格
            if sheet.cell(row=count, column=i).value ==sheet.cell(row=start_count, column=i).value:
                end_count=count
                continue
            if end_count > start_count:
                sheet.merge_cells(start_row=start_count, start_column=i, end_row=end_count, end_column=i)
                sheet.cell(row=start_count, column=i).alignment = Alignment(horizontal='center', vertical='center')
                print("merge(row%s,column%s) to (row%s,column%s)" %(start_count, i, end_count, i))
                start_count=count
                continue
            start_count=count
    wb.save(excelname)

def select_file():
    root=tk.Tk()
    root.withdraw()
    filepath=filedialog.askopenfilename()  #获取文件名
    return filepath

def maintest(excelname):
    file_name=select_file()
    out = xmind_to_dict(file_name)
    # excelname = filename.split('/')[-1].split('.')[0] + '.xls'
    xmind_cat(out[0]['topic']['topics'], excelname)

if __name__ == '__main__':
    maintest('case.xlsx')
