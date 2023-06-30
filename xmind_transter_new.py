import openpyxl
from openpyxl.styles import Alignment
from xmindparser import xmind_to_dict
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import re

"""适用于禅道的用例导入，xmind形式编写用例后，自动生成禅道导入格式的用例"""
def resolve_path(d, lists, title):
    title = title.strip()
    if len(title) == 0:
        concat_title = d['title'].strip()
    else:
        if 'makers' in d.keys():
            level = ''
            if 'priority-2' in d['makers']:
                level = 'P1'
            elif 'priority-1' in d['makers']:
                level = 'P0'
            elif 'priority-3' in d['makers']:
                level = 'P2'
            concat_title = title + '\t' + d['title'].strip() + '\t' + level
        else:
            concat_title = title + '\t' + d['title'].strip()
    if 'topics' not in d:
        lists.append(concat_title)
    else:
        for sub_d in d['topics']:
            resolve_path(sub_d, lists, concat_title)

def clean_precondition(precondition, content):
    # 定义正则表达式模式，匹配前置条件、步骤、预期的不同前缀
    pattern = r"^%s[:：；.]?\s*" % content
    # 使用正则表达式替换，将不同前缀转换为统一的格式
    cleaned_precondition = re.sub(pattern, "", precondition)
    # print(cleaned_precondition)
    return cleaned_precondition

def xmind_cat(lst):
    # wb = openpyxl.load_workbook(excelname)
    wb = openpyxl.Workbook()
    sheetname = wb.sheetnames
    sheet = wb[sheetname[0]]
    title_list = ['所属模块', '用例标题', '优先级', '前置条件', '步骤', '预期', '用例类型']
    for i in range(1, len(title_list) + 1):
        sheet.cell(row=1, column=i).value = title_list[i - 1]
    index = 1
    for h in range(len(lst)):
        lists = []
        resolve_path(lst[h], lists, '')
        prev_module_details = ''
        for j in range(len(lists)):
            lists[j] = lists[j].split('\t')
            # print(lists[j])
            start_column = 1
            if 6 > len(lists[j]) >= 4:
                module_details = '-'.join(lists[j][1:4])
                sheet.cell(row=j + index + 1, column=1).value = lists[j][0]
                sheet.cell(row=j + index + 1, column=2).value = module_details
                if len(lists[j]) > 4:
                    sheet.cell(row=j + index + 1, column=3).value = lists[j][4]
                sheet.cell(row=j + index + 1, column=7).value = "功能测试"
            elif len(lists[j]) == 6:
                # print(lists[j])
                # cond = lists[j][5].replace
                module_details = '-'.join(lists[j][1:4])
                if module_details == prev_module_details:
                    index -= 1
                    if lists[j][5].startswith('前置条件'):
                        sub_preconditions = clean_precondition(lists[j][5], "前置条件")
                        sheet.cell(row=j + index + 1, column=4).value = sub_preconditions

                    elif lists[j][5].startswith('步骤'):
                        sub_steps = clean_precondition(lists[j][5], "步骤")
                        sheet.cell(row=j + index + 1, column=5).value = sub_steps

                    elif lists[j][5].startswith('预期'):
                        sub_expected = clean_precondition(lists[j][5], "预期")
                        sheet.cell(row=j + index + 1, column=6).value = sub_expected

                else:
                    sheet.cell(row=j + index + 1, column=1).value = lists[j][0]
                    sheet.cell(row=j + index + 1, column=2).value = module_details
                    sheet.cell(row=j + index + 1, column=3).value = lists[j][4]
                    sheet.cell(row=j + index + 1, column=7).value = "功能测试"
                    if lists[j][5].startswith('前置条件'):
                        sub_preconditions = clean_precondition(lists[j][5], "前置条件")
                        sheet.cell(row=j + index + 1, column=4).value = sub_preconditions

                    elif lists[j][5].startswith('步骤'):
                        sub_steps = clean_precondition(lists[j][5], "步骤")
                        sheet.cell(row=j + index + 1, column=5).value = sub_steps

                    elif lists[j][5].startswith('预期'):
                        sub_expected = clean_precondition(lists[j][5], "预期")
                        sheet.cell(row=j + index + 1, column=6).value = sub_expected
                prev_module_details = module_details
        if j == len(lists) - 1:
            index += len(lists)
    now = datetime.now()
    date_time = now.strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"{date_time}_case.xlsx"
    wb.save(filename)


def select_file():
    filepath = filedialog.askopenfilename()
    return filepath


def maintest():
    file_name = select_file()
    out = xmind_to_dict(file_name)
    xmind_cat(out[0]['topic']['topics'])


if __name__ == '__main__':
    maintest()